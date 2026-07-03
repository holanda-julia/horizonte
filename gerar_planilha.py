# -*- coding: utf-8 -*-
"""
PROJETO HORIZONTE - GERADOR DE PLANILHA EXCEL
Desenvolvido para Júlia Aragão
"""

import sys
import os

# Garantir a instalação do openpyxl se necessário
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

def criar_planilha():
    wb = openpyxl.Workbook()
    # Remover aba padrão
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Definição de Cores (Paleta Luxury Verde & Dourado)
    fill_verde_escuro = PatternFill(start_color="092C22", end_color="092C22", fill_type="solid")
    fill_verde_medio = PatternFill(start_color="0F4C3A", end_color="0F4C3A", fill_type="solid")
    fill_verde_suave = PatternFill(start_color="EAF3F0", end_color="EAF3F0", fill_type="solid")
    fill_dourado = PatternFill(start_color="C29D66", end_color="C29D66", fill_type="solid")
    fill_cinza_claro = PatternFill(start_color="F2F5F4", end_color="F2F5F4", fill_type="solid")

    font_dourado_negrito = Font(name="Segoe UI", size=11, bold=True, color="C29D66")
    font_branco_negrito = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_padrao = Font(name="Segoe UI", size=10, color="2C3E35")
    font_negrito = Font(name="Segoe UI", size=10, bold=True, color="2C3E35")
    font_titulo = Font(name="Georgia", size=16, bold=True, color="092C22")
    font_subtitulo = Font(name="Segoe UI", size=10, italic=True, color="6B8276")

    align_centro = Alignment(horizontal="center", vertical="center")
    align_esquerda = Alignment(horizontal="left", vertical="center")
    align_direita = Alignment(horizontal="right", vertical="center")

    thin_border_side = Side(border_style="thin", color="E0E7E3")
    border_celula = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    border_total = Border(top=Side(border_style="thin", color="092C22"), bottom=Side(border_style="double", color="092C22"))

    # ==========================================
    # 1. ABA: FLUXO DE CAIXA
    # ==========================================
    ws_fluxo = wb.create_sheet("Fluxo de Caixa")
    ws_fluxo.views.sheetView[0].showGridLines = True

    # Cabeçalho da Aba
    ws_fluxo["A1"] = "PROJETO HORIZONTE — FLUXO DE CAIXA DETALHADO"
    ws_fluxo["A1"].font = font_titulo
    ws_fluxo["A2"] = "Planejamento financeiro de Júlia Aragão de Janeiro/2026 a Dezembro/2030 (Estimativas em R$)"
    ws_fluxo["A2"].font = font_subtitulo

    # Títulos da Tabela
    headers = [
        "Mês/Ano", "Salário Líquido", "Recebimento Carol", "13º Salário", "Natura/Extra", 
        "Gastos Fixos", "Gastos Esporádicos", "Dívidas Pessoas", "Parcelas Solange", 
        "Saldo Líquido Mês", "Saldo Acumulado"
    ]
    
    for col_idx, h in enumerate(headers, 1):
        cell = ws_fluxo.cell(row=4, column=col_idx, value=h)
        cell.font = font_branco_negrito
        cell.fill = fill_verde_escuro
        cell.alignment = align_centro
        cell.border = border_celula
        ws_fluxo.row_dimensions[4].height = 28

    # Gerar dados simulados mês a mês
    meses = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]
    row_count = 5
    
    for ano in range(2026, 2031):
        for m_idx, mes in enumerate(meses):
            mes_ano = f"{mes}/{ano}"
            
            # Condições de gastos fixos
            is_after_oct26 = (ano > 2026) or (ano == 2026 and m_idx >= 9)
            fixed_exp = 350.00 if is_after_oct26 else 450.00
            
            # Condição da Titia Solange
            is_solange = (ano == 2026 and m_idx >= 6) or (ano == 2027 and m_idx == 0)
            solange_val = 249.14 if is_solange else 0.00
            
            # Outros parcelamentos (começam em setembro/2026 simultaneamente)
            # Dívidas totais estimadas a serem parceladas:
            # Mila: 300, Vitinho: 250, Papis: 400, Chica: 350 (Total: 1300)
            # Parcelas mensais simultâneas de R$ 55 por pessoa. Total: R$ 220
            # Amortização simplificada na planilha
            is_debt_period = (ano > 2026) or (ano == 2026 and m_idx >= 8) # Set/2026
            
            # Controlando amortização simplificada por fórmulas Excel
            if is_debt_period:
                # Vamos colocar fórmulas ou valores que decrescem conforme os meses
                # Vitinho (250) acaba em 5 meses (Jan/2027)
                # Mila (300) acaba em 6 meses (Fev/2027)
                # Chica (350) acaba em 7 meses (Mar/2027)
                # Papis (400) acaba em 8 meses (Abr/2027)
                
                # Cálculo simples de parcelas ativas no mês
                active_debts = 0
                # Meses desde Set/26:
                meses_decorridos = (ano - 2026)*12 + (m_idx - 8)
                
                # Mila (6 parcelas de 50)
                if meses_decorridos < 6: active_debts += 55.00
                # Vitinho (5 parcelas de 50)
                if meses_decorridos < 5: active_debts += 55.00
                # Papis (8 parcelas de 50)
                if meses_decorridos < 8: active_debts += 55.00
                # Chica (7 parcelas de 50)
                if meses_decorridos < 7: active_debts += 55.00
                
                others_pay = active_debts
            else:
                others_pay = 0.00

            # 13º Salário
            thirteenth = 800.45 if m_idx in [10, 11] else 0.00
            
            # Escrever valores nas colunas correspondentes
            ws_fluxo.cell(row=row_count, column=1, value=mes_ano).alignment = align_centro
            ws_fluxo.cell(row=row_count, column=2, value=1600.90)
            ws_fluxo.cell(row=row_count, column=3, value=450.00)
            ws_fluxo.cell(row=row_count, column=4, value=thirteenth)
            ws_fluxo.cell(row=row_count, column=5, value=150.00)
            ws_fluxo.cell(row=row_count, column=6, value=fixed_exp)
            ws_fluxo.cell(row=row_count, column=7, value=200.00)
            ws_fluxo.cell(row=row_count, column=8, value=others_pay)
            ws_fluxo.cell(row=row_count, column=9, value=solange_val)
            
            # Fórmulas Excel para Saldo Líquido e Acumulado
            formula_net = f"=(B{row_count}+C{row_count}+D{row_count}+E{row_count})-(F{row_count}+G{row_count}+H{row_count}+I{row_count})"
            ws_fluxo.cell(row=row_count, column=10, value=formula_net)
            
            if row_count == 5:
                formula_acc = f"=J5"
            else:
                formula_acc = f"=K{row_count-1}+J{row_count}"
            ws_fluxo.cell(row=row_count, column=11, value=formula_acc)

            # Estilização das linhas de dados
            for col_idx in range(1, 12):
                cell = ws_fluxo.cell(row=row_count, column=col_idx)
                cell.font = font_padrao
                cell.border = border_celula
                if col_idx > 1:
                    cell.number_format = "R$ #,##0.00"
                    cell.alignment = align_direita
                if row_count % 2 == 0:
                    cell.fill = fill_cinza_claro

            ws_fluxo.row_dimensions[row_count].height = 20
            row_count += 1

    # Formatar Colunas automaticamente
    for col in ws_fluxo.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_fluxo.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # ==========================================
    # 2. ABA: PLANO DE QUITAÇÃO DE DÍVIDAS
    # ==========================================
    ws_dividas = wb.create_sheet("Plano de Dívidas")
    ws_dividas.views.sheetView[0].showGridLines = True

    ws_dividas["A1"] = "CONTROLE E CRONOGRAMA DE AMORTIZAÇÃO DE DÍVIDAS"
    ws_dividas["A1"].font = font_titulo
    ws_dividas["A2"] = "Acompanhamento detalhado das pessoas a quem Júlia deve"
    ws_dividas["A2"].font = font_subtitulo

    # Cabeçalho da Tabela
    headers_div = ["Membro", "Valor Inicial (R$)", "Parcela Acordada (R$)", "Data Início", "Meses Estimados", "Data Quitação"]
    for col_idx, h in enumerate(headers_div, 1):
        cell = ws_dividas.cell(row=4, column=col_idx, value=h)
        cell.font = font_branco_negrito
        cell.fill = fill_verde_medio
        cell.alignment = align_centro
        cell.border = border_celula
        ws_dividas.row_dimensions[4].height = 26

    dividas_data = [
        ["Titia Solange", 1743.98, 249.14, "Julho/2026", 7, "Janeiro/2027"],
        ["Mila", 300.00, 55.00, "Setembro/2026", 6, "Fevereiro/2027"],
        ["Vitinho", 250.00, 55.00, "Setembro/2026", 5, "Janeiro/2027"],
        ["Papis", 400.00, 55.00, "Setembro/2026", 8, "Abril/2027"],
        ["Madrinha Chica", 350.00, 55.00, "Setembro/2026", 7, "Março/2027"]
    ]

    for row_idx, r in enumerate(dividas_data, 5):
        for col_idx, val in enumerate(r, 1):
            cell = ws_dividas.cell(row=row_idx, column=col_idx, value=val)
            cell.font = font_padrao
            cell.border = border_celula
            if col_idx in [2, 3]:
                cell.number_format = "R$ #,##0.00"
                cell.alignment = align_direita
            elif col_idx in [1, 4, 6]:
                cell.alignment = align_esquerda
            else:
                cell.alignment = align_centro
        ws_dividas.row_dimensions[row_idx].height = 20

    # Adicionar totalizadores
    total_row = len(dividas_data) + 5
    ws_dividas.cell(row=total_row, column=1, value="TOTAL DE DÍVIDAS").font = font_negrito
    cell_total_val = ws_dividas.cell(row=total_row, column=2, value="=SUM(B5:B9)")
    cell_total_val.font = font_negrito
    cell_total_val.number_format = "R$ #,##0.00"
    cell_total_val.alignment = align_direita
    
    for c in range(1, 7):
        ws_dividas.cell(row=total_row, column=c).border = border_total

    # Ajuste colunas
    for col in ws_dividas.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_dividas.column_dimensions[col_letter].width = max(max_len + 4, 15)

    # ==========================================
    # 3. ABA: METAS (RESERVA E INTERCÂMBIO)
    # ==========================================
    ws_metas = wb.create_sheet("Metas & Investimentos")
    ws_metas.views.sheetView[0].showGridLines = True

    ws_metas["A1"] = "PLANEJAMENTO DE RESERVA E INTERCÂMBIO"
    ws_metas["A1"].font = font_titulo
    ws_metas["A2"] = "Simulação e custo estimado de intercâmbio em 6 opções de países e Reserva de Emergência"
    ws_metas["A2"].font = font_subtitulo

    # Reserva de Emergência Section
    ws_metas["A4"] = "1. RESERVA EMERGENCIAL DE SEGURANÇA"
    ws_metas["A4"].font = font_dourado_negrito
    
    headers_res = ["Meta Total", "Prazo Desejado", "Aporte Mensal Recom.", "Onde Investir"]
    for idx, h in enumerate(headers_res, 1):
        ws_metas.cell(row=5, column=idx, value=h).font = font_branco_negrito
        ws_metas.cell(row=5, column=idx).fill = fill_verde_medio
        ws_metas.cell(row=5, column=idx).alignment = align_centro
        ws_metas.cell(row=5, column=idx).border = border_celula
    
    ws_metas.cell(row=6, column=1, value=6000.00).number_format = "R$ #,##0.00"
    ws_metas.cell(row=6, column=2, value="12 a 18 meses")
    ws_metas.cell(row=6, column=3, value=150.00).number_format = "R$ #,##0.00"
    ws_metas.cell(row=6, column=4, value="Tesouro Selic ou CDB Liquidez Diária 100% CDI")
    
    for c in range(1, 5):
        cell = ws_metas.cell(row=6, column=c)
        cell.font = font_padrao
        cell.border = border_celula
        if c in [1, 3]:
            cell.alignment = align_direita
        else:
            cell.alignment = align_centro

    # Intercâmbio Section
    ws_metas["A9"] = "2. PROJETO INTERCÂMBIO (PRAZO DE 3 A 5 ANOS)"
    ws_metas["A9"].font = font_dourado_negrito

    headers_int = ["Destino", "Custo Médio Est. (R$)", "Duração", "Idioma Principal", "Exigência de Visto"]
    for idx, h in enumerate(headers_int, 1):
        ws_metas.cell(row=10, column=idx, value=h).font = font_branco_negrito
        ws_metas.cell(row=10, column=idx).fill = fill_verde_escuro
        ws_metas.cell(row=10, column=idx).alignment = align_centro
        ws_metas.cell(row=10, column=idx).border = border_celula
        
    int_data = [
        ["Malta 🇲🇹", 60000.00, "6 meses", "Inglês", "Fácil (Turista Schengen)"],
        ["Irlanda 🇮🇪", 75000.00, "8 meses (estudo+trabalho)", "Inglês", "Médio (Visto de Estudante local)"],
        ["Cabo Verde 🇨🇻", 35000.00, "3 meses", "Português / Crioulo", "Muito Fácil"],
        ["África do Sul 🇿🇦", 48000.00, "6 meses", "Inglês", "Fácil (Visto na chegada)"],
        ["Itália 🇮🇹", 72000.00, "6 meses", "Italiano", "Médio (Estudos)"],
        ["França 🇫🇷", 85000.00, "6 meses", "Francês", "Médio (Estudos / Cultural)"]
    ]

    for idx, r in enumerate(int_data, 11):
        for col_idx, val in enumerate(r, 1):
            cell = ws_metas.cell(row=idx, column=col_idx, value=val)
            cell.font = font_padrao
            cell.border = border_celula
            if col_idx == 2:
                cell.number_format = "R$ #,##0.00"
                cell.alignment = align_right = align_direita
            elif col_idx in [1, 4, 5]:
                cell.alignment = align_esquerda
            else:
                cell.alignment = align_centro
        ws_metas.row_dimensions[idx].height = 20

    # Autoajuste das colunas
    for col in ws_metas.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_metas.column_dimensions[col_letter].width = max(max_len + 4, 15)


    # ==========================================
    # 4. ABA: CONTROLE NATURA (RENDA EXTRA)
    # ==========================================
    ws_natura = wb.create_sheet("Renda Extra Natura")
    ws_natura.views.sheetView[0].showGridLines = True

    ws_natura["A1"] = "GERENCIAMENTO DE VENDAS NATURA"
    ws_natura["A1"].font = font_titulo
    ws_natura["A2"] = "Aba de lançamentos rápidos de produtos vendidos e comissão recebida"
    ws_natura["A2"].font = font_subtitulo

    # Títulos da Tabela de Pedidos
    headers_nat = ["Data", "Cliente", "Produto Vendido", "Valor da Venda (R$)", "Comissão Esperada (30%)", "Status Entrega"]
    for col_idx, h in enumerate(headers_nat, 1):
        cell = ws_natura.cell(row=4, column=col_idx, value=h)
        cell.font = font_branco_negrito
        cell.fill = fill_verde_medio
        cell.alignment = align_centro
        cell.border = border_celula
        ws_natura.row_dimensions[4].height = 26

    # Inserir alguns exemplos
    ex_natura = [
        ["2026-06-15", "Mariana Silva", "Perfume Kaiak", 149.90, "=D5*0.30", "Entregue"],
        ["2026-06-20", "Carlos Souza", "Creme TodoDia", 59.90, "=D6*0.30", "Entregue"],
        ["2026-07-01", "Ana Santos", "Desodorante Humor", 39.90, "=D7*0.30", "Pendente"]
    ]

    for row_idx, r in enumerate(ex_natura, 5):
        for col_idx, val in enumerate(r, 1):
            cell = ws_natura.cell(row=row_idx, column=col_idx, value=val)
            cell.font = font_padrao
            cell.border = border_celula
            if col_idx == 4:
                cell.number_format = "R$ #,##0.00"
                cell.alignment = align_direita
            elif col_idx == 5:
                cell.number_format = "R$ #,##0.00"
                cell.alignment = align_right = align_direita
                cell.font = font_negrito
            elif col_idx in [2, 3]:
                cell.alignment = align_esquerda
            else:
                cell.alignment = align_centro
        ws_natura.row_dimensions[row_idx].height = 20

    # Linha de Totais
    total_nat_row = len(ex_natura) + 5
    ws_natura.cell(row=total_nat_row, column=1, value="TOTAIS").font = font_negrito
    
    val_venda_total = ws_natura.cell(row=total_nat_row, column=4, value=f"=SUM(D5:D{total_nat_row-1})")
    val_venda_total.font = font_negrito
    val_venda_total.number_format = "R$ #,##0.00"
    val_venda_total.alignment = align_direita
    
    val_comissao_total = ws_natura.cell(row=total_nat_row, column=5, value=f"=SUM(E5:E{total_nat_row-1})")
    val_comissao_total.font = font_negrito
    val_comissao_total.number_format = "R$ #,##0.00"
    val_comissao_total.alignment = align_direita
    
    for c in range(1, 7):
        ws_natura.cell(row=total_nat_row, column=c).border = border_total

    # Ajuste colunas
    for col in ws_natura.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_natura.column_dimensions[col_letter].width = max(max_len + 4, 15)


    # Salvar o workbook
    filepath = "Projeto_Horizonte.xlsx"
    wb.save(filepath)
    print(f"Planilha gerada com sucesso em: {os.path.abspath(filepath)}")

if __name__ == "__main__":
    criar_planilha()
